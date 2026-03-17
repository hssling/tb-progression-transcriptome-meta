from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import re

from openpyxl import load_workbook
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
RAW_XLSX = ROOT / "data" / "raw" / "GSE79362" / "GSE79362_primarySampleJunctions.xlsx"
RAW_META = ROOT / "data" / "processed" / "GSE79362" / "metadata.parquet"
OUT_DIR = ROOT / "data" / "processed" / "GSE79362_genelevel"


def normalize_gene(gene: object) -> str:
    if gene is None:
        return ""
    g = str(gene).strip().upper()
    if g in {"", "NAN", "NA", "---"}:
        return ""
    return g


def parse_sample_fields(sample_id: str) -> tuple[str, str, int]:
    s = str(sample_id)
    if "DAY0" in s:
        return subject_id(s), "DAY0", 0
    if "DAY180" in s:
        return subject_id(s), "DAY180", 180
    if "DAY360" in s:
        return subject_id(s), "DAY360", 360
    if "DAY540" in s:
        return subject_id(s), "DAY540", 540
    if "IC" in s:
        return subject_id(s), "IC", 999
    return subject_id(s), "OTHER", 50


def subject_id(sample_id: str) -> str:
    s = str(sample_id)
    m = re.match(r"^(\d{2})_(\d{4})", s)
    if m:
        return f"{m.group(1)}{m.group(2)}"
    m = re.match(r"^(\d{6})", s)
    if m:
        return m.group(1)
    m = re.match(r"^(\d{1,3})_", s)
    if m:
        return m.group(1)
    return s.split("_")[0]


def aggregate_sheet(sheet) -> tuple[dict[str, dict[str, float]], set[str]]:
    header = [cell for cell in next(sheet.iter_rows(values_only=True))]
    gene_idx = header.index("gene")
    sample_cols = [idx for idx, name in enumerate(header) if name not in {"entry", "strand", "chr", "start", "end", "gene"}]
    sample_names = [header[idx] for idx in sample_cols]
    gene_sample = defaultdict(lambda: defaultdict(float))
    for row in sheet.iter_rows(values_only=True):
        gene = normalize_gene(row[gene_idx])
        if not gene:
            continue
        for idx, sample in zip(sample_cols, sample_names, strict=False):
            value = row[idx]
            if value is None:
                continue
            try:
                gene_sample[gene][sample] += float(value)
            except (TypeError, ValueError):
                continue
    return gene_sample, set(sample_names)


def build_expression() -> pd.DataFrame:
    wb = load_workbook(RAW_XLSX, read_only=True, data_only=True)
    merged: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    sample_names: set[str] = set()
    for sheet_name in wb.sheetnames:
        gene_sample, sheet_samples = aggregate_sheet(wb[sheet_name])
        sample_names.update(sheet_samples)
        for gene, sample_map in gene_sample.items():
            for sample, value in sample_map.items():
                merged[gene][sample] += value
    genes = sorted(merged)
    samples = sorted(sample_names)
    rows = []
    for gene in genes:
        row = {"gene": gene}
        row.update({sample: merged[gene].get(sample, 0.0) for sample in samples})
        rows.append(row)
    expr = pd.DataFrame(rows).set_index("gene").T.reset_index().rename(columns={"index": "sample_id"})
    return expr


def build_metadata(expr: pd.DataFrame) -> pd.DataFrame:
    meta = pd.read_parquet(RAW_META).copy()
    meta = meta[meta["sample_id"].isin(expr["sample_id"])].copy()
    parsed = meta["sample_id"].map(parse_sample_fields)
    meta["subject_id"] = parsed.map(lambda x: x[0])
    meta["timepoint_label"] = parsed.map(lambda x: x[1])
    meta["timepoint_month"] = parsed.map(lambda x: x[2] / 30 if x[2] < 900 else None)
    return meta


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    expr = build_expression()
    meta = build_metadata(expr)
    expr = expr[expr["sample_id"].isin(meta["sample_id"])].copy()
    expr.to_parquet(OUT_DIR / "expression.parquet", index=False)
    meta.to_parquet(OUT_DIR / "metadata.parquet", index=False)
    summary = pd.DataFrame(
        [
            {
                "cohort_id": "GSE79362_genelevel",
                "n_samples": len(meta),
                "n_subjects": meta["subject_id"].nunique(),
                "n_genes": expr.shape[1] - 1,
                "progressors": int(meta["progressor"].sum()),
                "non_progressors": int((1 - meta["progressor"]).sum()),
            }
        ]
    )
    summary.to_csv(OUT_DIR / "remap_summary.csv", index=False)
    print(f"Gene-level GSE79362 written to {OUT_DIR}")


if __name__ == "__main__":
    main()
